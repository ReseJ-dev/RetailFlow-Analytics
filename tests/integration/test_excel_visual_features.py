"""Structural checks for the polished Excel management report."""

from dataclasses import replace
from datetime import UTC, datetime
from pathlib import Path

import openpyxl

from retailflow.analytics import (
    calculate_inventory_metrics,
    calculate_returns_analytics,
    calculate_sales_analytics,
    generate_recommendations,
)
from retailflow.analytics.comparisons import compare_periods
from retailflow.analytics.models import ReturnsKPIs, SalesKPIs
from retailflow.reporting.excel_report import generate_excel_report

from .test_excel_report import _processing_result


def _report_inputs():
    processing = _processing_result()
    sales = calculate_sales_analytics(processing.processed_orders, processing.returns)
    returns = calculate_returns_analytics(processing.processed_orders, processing.returns)
    inventory = calculate_inventory_metrics(
        processing.inventory,
        processing.processed_orders,
        processing.returns,
        as_of_date="2025-01-31",
    )
    return processing, sales, returns, inventory


def test_management_report_contains_visual_and_navigation_features(tmp_path: Path) -> None:
    processing, sales, returns, inventory = _report_inputs()
    generated = generate_excel_report(
        processing,
        sales,
        returns,
        inventory_analytics=inventory,
        recommendations=generate_recommendations(inventory),
        previous_sales_analytics=sales,
        period_comparison=compare_periods(sales.kpis, sales.kpis),
        output_directory=tmp_path,
        filename="visual-report.xlsx",
        report_id="VISUAL-1",
        generated_at=datetime(2025, 2, 1, tzinfo=UTC),
    )

    workbook = openpyxl.load_workbook(generated.report_path)
    assert sum(len(sheet._charts) for sheet in workbook.worksheets) >= 7
    assert workbook["00_Cover"]["A15"].hyperlink is not None
    for name in workbook.sheetnames[1:]:
        sheet = workbook[name]
        if name != "01_Executive_Summary":
            assert sheet["A2"].hyperlink is not None
        assert sheet.print_area
    assert workbook["02_Sales_Analysis"].freeze_panes == "A5"
    assert workbook["07_Processed_Data"].tables
    inventory_rules = sum(
        len(item.rules) for item in workbook["04_Inventory"].conditional_formatting
    )
    quality_rules = sum(
        len(item.rules) for item in workbook["06_Data_Quality"].conditional_formatting
    )
    assert inventory_rules >= 6
    assert quality_rules >= 3


def test_empty_analytical_sections_show_messages_without_blank_charts(tmp_path: Path) -> None:
    processing, sales, returns, inventory = _report_inputs()
    empty_sales = replace(
        sales,
        kpis=SalesKPIs(),
        enriched_orders=sales.enriched_orders.iloc[0:0],
        daily_revenue=sales.daily_revenue.iloc[0:0],
        weekly_revenue=sales.weekly_revenue.iloc[0:0],
        category_performance=sales.category_performance.iloc[0:0],
        country_performance=sales.country_performance.iloc[0:0],
        channel_performance=sales.channel_performance.iloc[0:0],
        top_products_by_revenue=sales.top_products_by_revenue.iloc[0:0],
        top_products_by_gross_profit=sales.top_products_by_gross_profit.iloc[0:0],
    )
    empty_returns = replace(
        returns,
        kpis=ReturnsKPIs(),
        enriched_returns=returns.enriched_returns.iloc[0:0],
        return_reasons=returns.return_reasons.iloc[0:0],
        products_by_return_rate=returns.products_by_return_rate.iloc[0:0],
    )
    generated = generate_excel_report(
        processing,
        empty_sales,
        empty_returns,
        inventory_analytics=inventory.iloc[0:0],
        output_directory=tmp_path,
        filename="empty-report.xlsx",
    )

    workbook = openpyxl.load_workbook(generated.report_path)
    assert sum(len(sheet._charts) for sheet in workbook.worksheets) == 0
    return_values = [cell.value for row in workbook["05_Returns"] for cell in row]
    assert "No returns were recorded during the selected period." in return_values
