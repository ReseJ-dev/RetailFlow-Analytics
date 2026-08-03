"""Integration tests for the initial Excel management report."""

from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from retailflow.analytics import (
    calculate_inventory_metrics,
    calculate_returns_analytics,
    calculate_sales_analytics,
    generate_recommendations,
)
from retailflow.common.exceptions import ReportGenerationError
from retailflow.ingestion.models import FileMetadata
from retailflow.models import (
    DatasetProcessingStatistics,
    ProcessingResult,
    ProcessingStatistics,
)
from retailflow.reporting.excel_report import REQUIRED_WORKSHEETS, generate_excel_report
from retailflow.validation import DatasetType, ValidationIssue, ValidationSeverity


def _processing_result() -> ProcessingResult:
    """Create a small, manually verifiable processing result."""
    orders = pd.DataFrame(
        {
            "order_id": ["O-1", "O-2"],
            "order_date": [pd.Timestamp("2025-01-10"), pd.Timestamp("2025-01-11")],
            "product_id": ["P-1", "P-2"],
            "product_name": ["Desk", "Chair"],
            "category": ["Office", "Office"],
            "quantity": [2, 1],
            "unit_price": [10.0, 20.0],
            "discount": [0.0, 0.0],
            "purchase_cost": [5.0, 10.0],
            "currency": ["EUR", "EUR"],
            "country": ["Cyprus", "Germany"],
            "sales_channel": ["website", "amazon"],
            "order_status": ["completed", "completed"],
            "source_file": ["orders.csv", "orders.csv"],
            "source_row_number": [2, 3],
            "processing_status": ["processed", "processed"],
        }
    )
    products = pd.DataFrame(
        {
            "product_id": ["P-1", "P-2"],
            "product_name": ["Desk", "Chair"],
            "category": ["Office", "Office"],
            "purchase_cost": [5.0, 10.0],
            "recommended_price": [10.0, 20.0],
        }
    )
    inventory = pd.DataFrame(
        {
            "product_id": ["P-1", "P-2"],
            "product_name": ["Desk", "Chair"],
            "warehouse": ["Nicosia", "Berlin"],
            "stock_quantity": [3, 100],
            "reserved_quantity": [1, 0],
            "reorder_level": [5, 10],
            "last_restock_date": [pd.Timestamp("2025-01-01"), pd.Timestamp("2025-01-02")],
        }
    )
    returns = pd.DataFrame(
        {
            "return_id": ["R-1"],
            "order_id": ["O-1"],
            "product_id": ["P-1"],
            "return_date": [pd.Timestamp("2025-01-12")],
            "quantity": [1],
            "return_reason": ["Defective product"],
            "refund_amount": [10.0],
        }
    )
    targets = pd.DataFrame(
        {
            "month": ["2025-01"],
            "revenue_target": [1000.0],
            "profit_target": [300.0],
            "orders_target": [50],
        }
    )
    excluded = pd.DataFrame(
        {
            "source_dataset": ["orders"],
            "order_id": ["O-BAD"],
            "processing_status": ["excluded"],
            "exclusion_reason": ["invalid_quantity"],
        }
    )
    issue = ValidationIssue(
        ValidationSeverity.ERROR,
        DatasetType.ORDERS,
        "orders.csv",
        4,
        "quantity",
        "invalid_quantity",
        "Quantity must be a whole number.",
        "many",
        "Provide a positive whole number.",
        False,
    )
    row_counts = {
        DatasetType.ORDERS: (3, 2, 1, 1),
        DatasetType.PRODUCTS: (2, 2, 0, 0),
        DatasetType.INVENTORY: (2, 2, 0, 0),
        DatasetType.RETURNS: (1, 1, 0, 0),
        DatasetType.MONTHLY_TARGETS: (1, 1, 0, 0),
    }
    statistics = ProcessingStatistics(
        {
            dataset_type: DatasetProcessingStatistics(*counts)
            for dataset_type, counts in row_counts.items()
        }
    )
    frames = {
        DatasetType.ORDERS: orders,
        DatasetType.PRODUCTS: products,
        DatasetType.INVENTORY: inventory,
        DatasetType.RETURNS: returns,
        DatasetType.MONTHLY_TARGETS: targets,
    }
    metadata = {
        dataset_type: FileMetadata(
            filename=f"{dataset_type.value}.csv",
            file_type="csv",
            file_size=100,
            row_count=len(frame),
            column_count=len(frame.columns),
            columns=tuple(str(column) for column in frame.columns),
            detected_delimiter=",",
            detected_encoding="utf-8",
        )
        for dataset_type, frame in frames.items()
    }
    return ProcessingResult(
        processed_orders=orders,
        products=products,
        inventory=inventory,
        returns=returns,
        targets=targets,
        excluded_rows=excluded,
        validation_issues=(issue,),
        statistics=statistics,
        source_metadata=metadata,
    )


def test_generate_excel_report_creates_every_required_worksheet(tmp_path: Path) -> None:
    """The report should be non-empty and contain important summary values."""
    processing = _processing_result()
    sales = calculate_sales_analytics(processing.processed_orders, processing.returns)
    returns = calculate_returns_analytics(processing.processed_orders, processing.returns)
    inventory = calculate_inventory_metrics(
        processing.inventory,
        processing.processed_orders,
        processing.returns,
        as_of_date="2025-01-31",
    )
    recommendations = generate_recommendations(inventory)

    generated = generate_excel_report(
        processing,
        sales,
        returns,
        inventory_analytics=inventory,
        recommendations=recommendations,
        output_directory=tmp_path / "nested" / "reports",
        filename="management_report.xlsx",
        company_name="Northstar Retail Group",
        default_currency="EUR",
        report_id="REPORT-TEST-001",
        generated_at=datetime(2025, 2, 1, 12, 30, tzinfo=UTC),
    )

    assert generated.report_path.exists()
    assert generated.file_size > 0
    assert generated.statistics.worksheet_count == 9
    assert generated.statistics.processed_order_rows == 2

    workbook = openpyxl.load_workbook(generated.report_path, data_only=False)
    assert workbook.sheetnames == list(REQUIRED_WORKSHEETS)
    assert workbook["00_Cover"]["A1"].value == "RetailFlow Analytics Management Report"
    assert workbook["00_Cover"]["B4"].value == "Northstar Retail Group"
    assert workbook["00_Cover"]["B5"].value == "REPORT-TEST-001"
    assert workbook["01_Executive_Summary"]["A1"].value == "Executive Summary"
    assert workbook["01_Executive_Summary"]["B5"].value == 40
    assert "EUR" in workbook["01_Executive_Summary"]["B5"].number_format
    assert workbook["08_Report_Metadata"]["B3"].value == "REPORT-TEST-001"
    assert workbook["07_Processed_Data"].tables


def test_report_does_not_overwrite_without_explicit_permission(tmp_path: Path) -> None:
    """Existing report files must remain protected by default."""
    processing = _processing_result()
    sales = calculate_sales_analytics(processing.processed_orders, processing.returns)
    returns = calculate_returns_analytics(processing.processed_orders, processing.returns)
    report_path = tmp_path / "existing.xlsx"
    report_path.write_bytes(b"existing workbook placeholder")

    with pytest.raises(ReportGenerationError, match="report file already exists"):
        generate_excel_report(
            processing,
            sales,
            returns,
            output_directory=tmp_path,
            filename=report_path.name,
        )

    assert report_path.read_bytes() == b"existing workbook placeholder"

    overwritten = generate_excel_report(
        processing,
        sales,
        returns,
        output_directory=tmp_path,
        filename=report_path.name,
        overwrite=True,
    )
    assert overwritten.file_size > len(b"existing workbook placeholder")
    assert openpyxl.load_workbook(overwritten.report_path).sheetnames == list(REQUIRED_WORKSHEETS)
